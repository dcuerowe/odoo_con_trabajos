"""Barrido histórico del formulario a un Excel de prueba LOCAL.

Descarga todas las submissions desde una fecha, las procesa con el pipeline
normal (`ordenar_respuestas` -> `process_entrys`) y escribe un `.xlsx` local con
las tres tablas: Terreno, Inspección y Residuos.

Es deliberadamente de solo lectura respecto a producción:

  - NO escribe en SharePoint (no usa `excel_manager.send_data`).
  - NO consulta ni modifica `form_entries.db` (no usa `check_new_sub`), así que
    correrlo no marca OTs como procesadas ni bloquea la corrida automática.
  - NO genera ni sube informes PDF (no usa `adjuntar_informes`).

A diferencia de la escritura a SharePoint, acá se vuelcan TODAS las columnas de
cada DataFrame, no solo las que calzan con un encabezado existente: la idea es
poder revisar qué trae el pipeline, incluido lo que hoy no tiene destino.

Uso (desde el directorio `trabajos_terreno/`):

    python barrido_prueba.py
    python barrido_prueba.py --desde 2025-11-01 --hasta 2026-08-21
    python barrido_prueba.py --salida ~/Desktop/barrido.xlsx
    python barrido_prueba.py --cache-json crudo.json    # guarda la descarga
    python barrido_prueba.py --cache-json crudo.json    # 2a vez: la reutiliza

Requiere `CONNECTEAM_API_KEY` en el `.env` (o en el entorno). No requiere
credenciales de SharePoint.
"""
import argparse
import json
import os
from datetime import date, datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import CONNECTEAM_API_KEY
from connecteam_api import form_structure, submissions_en_rango
from data_processing import ordenar_respuestas
from processor import process_entrys

# Fecha desde la que se quiere trazabilidad de residuos.
DESDE_POR_DEFECTO = date(2025, 11, 1)

# (hoja, nombre de tabla) por DataFrame, alineado con Terreno.xlsx.
HOJAS = [
    ('Terreno', 'OTS'),
    ('Inspección', 'Ronda'),
    ('Residuos', 'Residuos'),
]


def _fecha(texto):
    return datetime.strptime(texto, '%Y-%m-%d').date()


def _limpiar_para_excel(df):
    """Deja el DataFrame escribible por openpyxl.

    - Descarta `_fotos`, la columna auxiliar de listas de URLs que normalmente
      consume y elimina `report_manager.adjuntar_informes` (acá no se llama).
    - Convierte cualquier celda con lista/tupla/dict a texto: openpyxl no puede
      escribir esos tipos y aborta la hoja completa.
    """
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.drop(columns=[c for c in ('_fotos',) if c in df.columns])

    def plano(v):
        if isinstance(v, (list, tuple)):
            return ', '.join(str(x) for x in v)
        if isinstance(v, dict):
            return json.dumps(v, ensure_ascii=False)
        return v

    return df.map(plano)


VISITA_RONDA = '(R) Ronda diaria de Inspección'


def _procesar(ordenadas):
    """Procesa el barrido aislando las OTs que fallan, sin inventar fallas.

    En producción `process_entrys` recibe el lote completo, así que una sola OT
    con una pregunta ausente aborta todo el lote: varias ramas de processor.py
    indexan columnas directo, sin `.get`. En un barrido de meses eso es casi
    seguro, y para revisar datos conviene quedarse con todo lo que sí procesa
    más una lista explícita de lo que no.

    Las rondas de inspección van en **un solo lote**, no OT por OT: son un
    passthrough que depende de la unión de columnas del `pd.concat` aguas abajo
    (`processor.py` hace `drop(columns=['Fotos '])` sin proteger). Aisladas, una
    ronda sin fotos falla con KeyError aunque en producción funcione — sería una
    falla inventada por el propio barrido. Los trabajos en terreno, en cambio,
    sí se procesan de a uno: ahí vive el indexado frágil por instancia.
    """
    es_ronda = ordenadas['Tipo de visita realizada'] == VISITA_RONDA
    rondas, trabajos = ordenadas[es_ronda], ordenadas[~es_ronda]

    terrenos, inspecciones, residuos_list, fallidas = [], [], [], []

    def acumular(t, i, r):
        for destino, parcial in ((terrenos, t), (inspecciones, i),
                                 (residuos_list, r)):
            if parcial is not None and not parcial.empty:
                destino.append(parcial)

    def anotar_falla(fila, exc):
        fallidas.append({
            'OT': fila.get('#'),
            'Tipo de visita': fila.get('Tipo de visita realizada'),
            'Error': f'{type(exc).__name__}: {exc}',
        })

    if not rondas.empty:
        try:
            acumular(*process_entrys(rondas, CONNECTEAM_API_KEY))
        except Exception as exc:
            print(f'  El lote de rondas falló ({exc}); se reintenta de a una.')
            for _, fila in rondas.iterrows():
                try:
                    acumular(*process_entrys(pd.DataFrame([fila]),
                                             CONNECTEAM_API_KEY))
                except Exception as exc_fila:
                    anotar_falla(fila, exc_fila)

    for _, fila in trabajos.iterrows():
        try:
            acumular(*process_entrys(pd.DataFrame([fila]), CONNECTEAM_API_KEY))
        except Exception as exc:
            anotar_falla(fila, exc)

    def unir(partes):
        return (pd.concat(partes, ignore_index=True) if partes
                else pd.DataFrame())

    if fallidas:
        print(f'  AVISO: {len(fallidas)} de {len(ordenadas)} OTs no se pudieron '
              f'procesar (ver hoja "OTs con error").')

    return (unir(terrenos), unir(inspecciones), unir(residuos_list),
            pd.DataFrame(fallidas))


def _escribir_hoja(wb, df, sheet_name, table_name, sin_tablas=False):
    ws = wb.create_sheet(sheet_name)

    if df.empty:
        ws['A1'] = f'Sin datos para {sheet_name} en el rango consultado.'
        return 0

    ws.append([str(c) for c in df.columns])
    for fila in df.itertuples(index=False, name=None):
        ws.append([None if pd.isna(v) else v for v in fila])

    ref = f'A1:{get_column_letter(len(df.columns))}{len(df) + 1}'
    if sin_tablas:
        # Solo autofiltro: una tabla nombrada de más es lo primero que Excel
        # "repara" si algo no le calza, y en un archivo de revisión no aporta.
        ws.auto_filter.ref = ref
    else:
        tabla = Table(displayName=table_name, ref=ref)
        tabla.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium2', showRowStripes=True)
        ws.add_table(tabla)

    for idx, columna in enumerate(df.columns, start=1):
        ancho = max(len(str(columna)), 12)
        ws.column_dimensions[get_column_letter(idx)].width = min(ancho + 2, 45)
    ws.freeze_panes = 'A2'
    return len(df)


def _hoja_resumen(wb, resumen, desde, hasta, n_submissions):
    ws = wb.create_sheet('Resumen', 0)
    ws['A1'] = 'Barrido de prueba — NO es el modelo de datos de producción'
    filas = [
        ('Generado', datetime.now().strftime('%d-%m-%Y %H:%M')),
        ('Rango consultado', f'{desde:%d-%m-%Y} a {hasta:%d-%m-%Y}'),
        ('Submissions descargadas', n_submissions),
        ('', ''),
    ]
    filas += [(f'Filas en {hoja}', n) for hoja, n in resumen.items()]
    for i, (etiqueta, valor) in enumerate(filas, start=3):
        ws.cell(row=i, column=1, value=etiqueta)
        ws.cell(row=i, column=2, value=valor)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 34


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--desde', type=_fecha, default=DESDE_POR_DEFECTO,
                        help='Fecha inicial YYYY-MM-DD (default 2025-11-01).')
    parser.add_argument('--hasta', type=_fecha, default=None,
                        help='Fecha final YYYY-MM-DD (default: hoy).')
    parser.add_argument('--salida', default='barrido_prueba.xlsx',
                        help='Ruta del .xlsx a generar.')
    parser.add_argument('--cache-json', default=None,
                        help='Archivo donde guardar/reusar la descarga cruda, '
                             'para reprocesar sin volver a golpear la API.')
    parser.add_argument('--sin-tablas', action='store_true',
                        help='No crear tablas nombradas, solo autofiltro. Útil '
                             'si Excel se queja del archivo al abrirlo.')
    parser.add_argument('--csv-dir', default=None,
                        help='Además del xlsx, volcar un CSV por hoja en este '
                             'directorio (UTF-8 con BOM, separador ;).')
    args = parser.parse_args()

    hasta = args.hasta or date.today()

    if args.cache_json and os.path.exists(args.cache_json):
        print(f'Reutilizando descarga cacheada: {args.cache_json}')
        with open(args.cache_json, encoding='utf-8') as fh:
            crudo = json.load(fh)
        estructura, respuestas = crudo['estructura'], crudo['respuestas']
    else:
        if not CONNECTEAM_API_KEY:
            raise SystemExit(
                'Falta CONNECTEAM_API_KEY. Definila en el .env de '
                'trabajos_terreno/ y volvé a correr.'
            )
        print('Descargando estructura del formulario...')
        estructura = form_structure(CONNECTEAM_API_KEY)
        print(f'Descargando submissions {args.desde} -> {hasta}...')
        respuestas = submissions_en_rango(
            CONNECTEAM_API_KEY, args.desde, hasta)
        if args.cache_json:
            with open(args.cache_json, 'w', encoding='utf-8') as fh:
                json.dump({'estructura': estructura, 'respuestas': respuestas},
                          fh, ensure_ascii=False)
            print(f'Descarga guardada en {args.cache_json}')

    n_submissions = len(respuestas.get('data', {}).get('formSubmissions', []))
    print(f'\nSubmissions a procesar: {n_submissions}')
    if not n_submissions:
        raise SystemExit('El rango no devolvió submissions. Nada que procesar.')

    print('Aplanando respuestas...')
    ordenadas = ordenar_respuestas(estructura, respuestas)
    print(f'  OTs: {len(ordenadas)} | columnas: {len(ordenadas.columns)}')

    print('Procesando (esto resuelve el nombre de cada técnico vía API)...')
    terreno, inspeccion, residuos, fallidas = _procesar(ordenadas)

    dfs = [_limpiar_para_excel(d) for d in (terreno, inspeccion, residuos)]

    wb = Workbook()
    wb.remove(wb.active)
    resumen = {}
    for df, (hoja, tabla) in zip(dfs, HOJAS):
        resumen[hoja] = _escribir_hoja(wb, df, hoja, tabla, args.sin_tablas)
    resumen['OTs con error'] = _escribir_hoja(
        wb, fallidas, 'OTs con error', 'OTsConError', args.sin_tablas)
    _hoja_resumen(wb, resumen, args.desde, hasta, n_submissions)
    wb.save(args.salida)

    print('\n--- Resumen ---')
    for hoja, n in resumen.items():
        print(f'  {hoja}: {n} fila(s)')
    if resumen.get('Residuos'):
        print('\nResiduos por categoría:')
        print(dfs[2]['Categoría'].value_counts(dropna=False).to_string())
    if args.csv_dir:
        os.makedirs(args.csv_dir, exist_ok=True)
        nombres = [h for h, _ in HOJAS] + ['OTs con error']
        for df, hoja in zip(dfs + [fallidas], nombres):
            if df.empty:
                continue
            ruta = os.path.join(args.csv_dir, f'{hoja}.csv')
            # BOM + ';' para que Excel en es-CL lo abra en columnas al doble clic.
            df.to_csv(ruta, index=False, sep=';', encoding='utf-8-sig')
            print(f'  CSV: {ruta}')

    if not fallidas.empty:
        print(f'\nOTs con error ({len(fallidas)}):')
        print(fallidas['Error'].value_counts().to_string())
    print(f'\nExcel generado: {os.path.abspath(args.salida)}')
    print('No se escribió nada en SharePoint ni en form_entries.db.')


if __name__ == '__main__':
    main()
