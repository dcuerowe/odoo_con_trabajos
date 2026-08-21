import io
from datetime import date, datetime
import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from config import EXCEL_URL


def _asegurar_hoja_y_tabla(wl, sheet_name, table_name, headers):
    """Crea la hoja y la tabla nombrada si no existen. Idempotente.

    Se necesita para tablas nuevas (p.ej. 'Residuos'): sin esto,
    `wl[sheet_name]` o `wh.tables[table_name]` lanzan KeyError, el except
    genérico de modify_excel_file lo convierte en un print y no se escribe nada.

    La tabla se registra con un objeto Table completo, de modo que openpyxl
    derive sus `tableColumns` de la fila de encabezados y Excel no marque el
    archivo como dañado.
    """
    if sheet_name in wl.sheetnames:
        wh = wl[sheet_name]
    else:
        wh = wl.create_sheet(sheet_name)
        print(f"[{table_name}] Hoja creada: {sheet_name!r}")

    if table_name in wh.tables:
        return

    # La hoja puede existir sin tabla nombrada (creada a mano). Solo se escriben
    # los encabezados si la fila 1 está vacía; si ya trae algo distinto, se
    # aborta en vez de pisar el trabajo de alguien. El error lo captura el
    # except de modify_excel_file, así que la corrida no escribe nada.
    presentes = [str(c.value).strip() for c in wh[1] if c.value is not None]
    esperados = [str(h).strip() for h in headers]
    if presentes and presentes != esperados:
        raise ValueError(
            f"La hoja {sheet_name!r} ya tiene encabezados que no coinciden con "
            f"los esperados y no existe la tabla {table_name!r}. "
            f"Encontrados: {presentes}. Esperados: {esperados}. "
            "Ajustalos a mano o dejá la fila 1 vacía para que el código la cree."
        )
    if not presentes:
        for idx, nombre_col in enumerate(headers, start=1):
            wh.cell(row=1, column=idx, value=nombre_col)

    ref = f"A1:{get_column_letter(len(headers))}1"
    tabla = Table(displayName=table_name, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    wh.add_table(tabla)
    print(f"[{table_name}] Tabla creada en {sheet_name!r} con ref {ref}")


def modify_excel_file(df, sheet_name, table_name, sharepoint_client, headers_si_falta=None):
    # Descarga el archivo Excel desde SharePoint
    excel = sharepoint_client.download_file(EXCEL_URL)
    if excel:
        try:
            # Convierte los bytes descargados en un objeto BytesIO para manipulación en memoria
            excel_file = io.BytesIO(excel)
            # Carga el archivo Excel en openpyxl
            wl = openpyxl.load_workbook(excel_file)

            # Para tablas nuevas: crea hoja + tabla nombrada si aún no existen.
            if headers_si_falta:
                _asegurar_hoja_y_tabla(wl, sheet_name, table_name, headers_si_falta)

            # Selecciona la hoja de trabajo especificada
            wh = wl[sheet_name]

            # Obtiene la tabla de la hoja por su nombre
            tabla = wh.tables[table_name]
            # Obtiene la referencia actual de la tabla (ejemplo: 'A1:H10')
            ref_actual = tabla.ref
            # Extrae las coordenadas inicial y final de la tabla
            referencia_inicial = ref_actual.split(':')[0]
            coordenada_final = ref_actual.split(':')[-1]
            fila_header, col_inicio = coordinate_to_tuple(referencia_inicial)
            fila_final_actual, columna_final_num = coordinate_to_tuple(coordenada_final)

            # --- Mapeo por NOMBRE de columna (no por posición) ---
            # Leemos los encabezados reales de la tabla en Excel y construimos
            # un mapa {nombre_encabezado: índice de columna}. Así cada valor se
            # escribe en la columna cuyo encabezado coincide con el del DataFrame,
            # sin depender del orden de columnas del DataFrame (que es variable,
            # sobre todo en Inspección por el dropna() + concat aguas arriba).
            header_a_columna = {}
            for col in range(col_inicio, columna_final_num + 1):
                encabezado = wh.cell(row=fila_header, column=col).value
                if encabezado is not None:
                    header_a_columna[str(encabezado).strip()] = col

            # Determinamos qué columnas del DataFrame calzan con un encabezado.
            columnas_df = [str(c).strip() for c in df.columns]
            columnas_match = [c for c in columnas_df if c in header_a_columna]
            columnas_sin_match = [c for c in columnas_df if c not in header_a_columna]

            # Columnas nuevas que el código sí debe crear en la tabla si faltan
            # (p.ej. 'Informe'). El resto de las no coincidentes se siguen
            # omitiendo con aviso, para no enmascarar errores de nombres.
            NUEVAS_PERMITIDAS = {'Informe'}
            nuevas = [c for c in columnas_sin_match if c in NUEVAS_PERMITIDAS]
            for nombre_col in nuevas:
                columna_final_num += 1
                wh.cell(row=fila_header, column=columna_final_num, value=nombre_col)
                header_a_columna[nombre_col] = columna_final_num
                columnas_match.append(nombre_col)
                print(f"[{table_name}] Columna nueva agregada a la tabla: {nombre_col!r}")

            columnas_sin_match = [c for c in columnas_sin_match if c not in NUEVAS_PERMITIDAS]
            if columnas_sin_match:
                print(
                    f"[{table_name}] Advertencia: columnas del DataFrame sin "
                    f"encabezado correspondiente en la tabla (se omiten): "
                    f"{columnas_sin_match}"
                )

            if not columnas_match:
                print(
                    f"[{table_name}] Ninguna columna del DataFrame coincide con "
                    f"los encabezados de la tabla. No se escribe nada."
                )
                return

            # Nuevos registros se insertan en la primera fila de datos (debajo del header)
            # para que los más recientes queden arriba.
            fila_inicio_nuevos_datos = fila_header + 1
            n_nuevas = len(df)

            # Hace espacio desplazando hacia abajo los datos existentes
            wh.insert_rows(idx=fila_inicio_nuevos_datos, amount=n_nuevas)

            # Escribe los nuevos datos en las filas recién insertadas, ubicando
            # cada valor en la columna de su encabezado.
            # Las claves de to_dict son los nombres ORIGINALES, pero
            # columnas_match viene de nombres ya strippeados. Sin normalizar, una
            # columna con espacios en los extremos ('Fecha visita ') provoca
            # KeyError, el except de abajo lo convierte en un print y se pierde
            # el lote completo sin subir nada, con las OT ya marcadas como
            # procesadas en form_entries.db (o sea, sin reintento).
            registros = df.rename(
                columns={c: str(c).strip() for c in df.columns}
            ).to_dict(orient='records')
            for i, fila_nueva in enumerate(registros):
                for nombre_col in columnas_match:
                    valor = fila_nueva[nombre_col]
                    columna_destino = header_a_columna[nombre_col]
                    cell = wh.cell(
                        row=fila_inicio_nuevos_datos + i,
                        column=columna_destino,
                        value=valor,
                    )
                    if isinstance(valor, (date, datetime)):
                        cell.number_format = 'DD/MM/YY'

            # Actualiza la referencia de la tabla para incluir las nuevas filas
            fila_final_nueva = fila_final_actual + n_nuevas
            columna_final_letra = get_column_letter(columna_final_num)
            nueva_referencia = f'{referencia_inicial}:{columna_final_letra}{fila_final_nueva}'
            tabla.ref = nueva_referencia
            # --- Fin del código de openpyxl ---

            # Guarda el archivo modificado en un nuevo stream de bytes
            excel_stream_out = io.BytesIO()
            wl.save(excel_stream_out)
            excel_stream_out.seek(0)  # Mueve el cursor al inicio del stream

            # Sube el archivo modificado de vuelta a SharePoint
            success = sharepoint_client.upload_file(EXCEL_URL, excel_stream_out)

        except Exception as e:
            print(f"Error al procesar el archivo Excel: {e}")

    else:
        print("No se pudo descargar el archivo.")

def send_data(df, sheet, table, sharepoint_client, headers_si_falta=None):
    # Enviamos el DataFrame completo (con nombres de columna) para que la
    # escritura en Excel se haga por nombre y no por posición.
    if df is not None and not df.empty:
        modify_excel_file(df, sheet, table, sharepoint_client, headers_si_falta)
