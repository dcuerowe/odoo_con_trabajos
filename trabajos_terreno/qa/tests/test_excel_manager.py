"""Casos TC-EX — excel_manager.modify_excel_file / send_data.

Se usa un libro Excel local en memoria y un cliente SharePoint simulado; nunca se accede
al Terreno.xlsx de producción.
"""
import io

import openpyxl
from openpyxl.worksheet.table import Table
import pytest

import excel_manager


class FakeSharepoint:
    """Cliente SharePoint simulado: descarga desde memoria y captura la subida."""

    def __init__(self, data: bytes):
        self._data = data
        self.uploaded = None

    def download_file(self, url):
        return self._data

    def upload_file(self, url, content_stream, content_type=None, folder_name=""):
        content_stream.seek(0)
        self.uploaded = content_stream.read()
        return True


def _workbook_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Terreno"
    ws.append(["Col1", "Col2"])      # cabecera (fila 1)
    ws.append(["a", "b"])            # fila de datos existente (fila 2)
    ws.add_table(Table(displayName="OTS", ref="A1:B2"))
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_tc_ex_01_insercion_incremental():
    """TC-EX-01: las filas nuevas se insertan al inicio (debajo del header) y la
    referencia de la tabla se expande.

    El comportamiento vigente inserta los registros nuevos en la primera fila de
    datos (más recientes arriba), desplazando hacia abajo los datos existentes.
    """
    sp = FakeSharepoint(_workbook_bytes())
    nuevas = [["c", "d"], ["e", "f"]]

    excel_manager.modify_excel_file(nuevas, "Terreno", "OTS", sp)

    assert sp.uploaded is not None
    wb = openpyxl.load_workbook(io.BytesIO(sp.uploaded))
    ws = wb["Terreno"]
    # nuevos registros arriba, debajo del header (fila 1)
    assert ws["A2"].value == "c"
    assert ws["A3"].value == "e"
    # la fila de datos previa se desplaza hacia abajo
    assert ws["A4"].value == "a"
    # referencia expandida de A1:B2 a A1:B4
    assert ws.tables["OTS"].ref == "A1:B4"


def test_tc_ex_03_dataframe_vacio_no_escribe():
    """TC-EX-03: send_data con DataFrame vacío no invoca la modificación."""
    import pandas as pd

    llamado = {"flag": False}

    def fake_modify(*args, **kwargs):
        llamado["flag"] = True

    original = excel_manager.modify_excel_file
    excel_manager.modify_excel_file = fake_modify
    try:
        excel_manager.send_data(pd.DataFrame(), "Terreno", "OTS", object())
    finally:
        excel_manager.modify_excel_file = original

    assert llamado["flag"] is False


def test_tc_ex_05_fallo_descarga(capsys):
    """TC-EX-05: si la descarga devuelve None, no se intenta subir."""
    class SPNulo:
        def download_file(self, url):
            return None
        def upload_file(self, *a, **k):
            raise AssertionError("no debe subirse si la descarga falló")

    excel_manager.modify_excel_file([["x", "y"]], "Terreno", "OTS", SPNulo())
    out = capsys.readouterr().out
    assert "No se pudo descargar" in out


_HEADERS_RESIDUOS_MIN = ["OT", "Categoría", "Cantidad"]


def test_tc_ex_04_crea_hoja_y_tabla_si_faltan():
    """TC-EX-04: con headers_si_falta se crea la hoja y la tabla nombrada.

    Sin esto, `wl[sheet_name]` lanza KeyError, el except genérico lo convierte en
    un print y no se escribe nada (caso de la tabla nueva 'Residuos').
    """
    import pandas as pd

    sp = FakeSharepoint(_workbook_bytes())   # el libro solo trae Terreno/OTS
    filas = pd.DataFrame([{"OT": "III-1", "Categoría": "Desechos", "Cantidad": "1"}])

    excel_manager.send_data(
        filas, "Residuos", "Residuos", sp,
        headers_si_falta=_HEADERS_RESIDUOS_MIN,
    )

    wb = openpyxl.load_workbook(io.BytesIO(sp.uploaded))
    assert "Residuos" in wb.sheetnames
    hoja = wb["Residuos"]
    tabla = hoja.tables["Residuos"]
    # Los tableColumns deben derivarse de los encabezados, o Excel marca el
    # archivo como dañado al abrirlo.
    assert [c.name for c in tabla.tableColumns] == _HEADERS_RESIDUOS_MIN
    assert tabla.ref == "A1:C2"
    assert [c.value for c in hoja[1]] == _HEADERS_RESIDUOS_MIN
    assert [c.value for c in hoja[2]] == ["III-1", "Desechos", "1"]
    # La tabla preexistente no se toca.
    assert wb["Terreno"].tables["OTS"].ref == "A1:B2"


def test_tc_ex_05_creacion_idempotente():
    """TC-EX-05: la segunda escritura reutiliza la tabla y solo agrega la fila."""
    import pandas as pd

    sp = FakeSharepoint(_workbook_bytes())
    filas = pd.DataFrame([{"OT": "III-1", "Categoría": "Desechos", "Cantidad": "1"}])

    excel_manager.send_data(
        filas, "Residuos", "Residuos", sp, headers_si_falta=_HEADERS_RESIDUOS_MIN)
    sp._data = sp.uploaded          # la siguiente corrida descarga lo ya subido
    excel_manager.send_data(
        filas, "Residuos", "Residuos", sp, headers_si_falta=_HEADERS_RESIDUOS_MIN)

    wb = openpyxl.load_workbook(io.BytesIO(sp.uploaded))
    assert wb.sheetnames.count("Residuos") == 1
    assert wb["Residuos"].tables["Residuos"].ref == "A1:C3"


def test_tc_ex_06_columna_con_espacios_no_pierde_el_lote():
    """TC-EX-06: una columna del DataFrame con espacios en los extremos calza con
    el encabezado sin ellos. Antes daba KeyError, el except lo silenciaba y se
    perdía el lote completo sin subir nada."""
    import pandas as pd

    sp = FakeSharepoint(_workbook_bytes())
    filas = pd.DataFrame([{"Col1 ": "v1", "Col2": "v2"}])

    excel_manager.modify_excel_file(filas, "Terreno", "OTS", sp)

    assert sp.uploaded is not None
    wb = openpyxl.load_workbook(io.BytesIO(sp.uploaded))
    assert [c.value for c in wb["Terreno"][2]] == ["v1", "v2"]


def test_tc_ex_07_no_pisa_encabezados_existentes():
    """TC-EX-07: si la hoja existe con encabezados distintos y sin tabla, se
    aborta sin escribir en vez de sobreescribir lo que puso alguien a mano."""
    import pandas as pd

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Terreno"
    ws.append(["Col1", "Col2"])
    ws.add_table(Table(displayName="OTS", ref="A1:A1"))
    hoja = wb.create_sheet("Residuos")
    hoja.append(["OT", "Otra cosa", "Cantidad"])     # encabezados a mano, distintos
    buf = io.BytesIO()
    wb.save(buf)

    sp = FakeSharepoint(buf.getvalue())
    filas = pd.DataFrame([{"OT": "III-1", "Categoría": "Desechos", "Cantidad": "1"}])

    excel_manager.send_data(
        filas, "Residuos", "Residuos", sp,
        headers_si_falta=_HEADERS_RESIDUOS_MIN,
    )

    assert sp.uploaded is None                       # no se subió nada
    wb2 = openpyxl.load_workbook(io.BytesIO(sp._data))
    assert [c.value for c in wb2["Residuos"][1]] == ["OT", "Otra cosa", "Cantidad"]


def test_tc_ex_08_reusa_hoja_vacia_preexistente():
    """TC-EX-08: una hoja que existe pero está vacía (el caso real de
    Terreno.xlsx) se reutiliza y se le crea la tabla, sin duplicar la hoja."""
    import pandas as pd

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Terreno"
    ws.append(["Col1", "Col2"])
    ws.add_table(Table(displayName="OTS", ref="A1:A1"))
    wb.create_sheet("Residuos")                      # vacía, sin tabla
    buf = io.BytesIO()
    wb.save(buf)

    sp = FakeSharepoint(buf.getvalue())
    filas = pd.DataFrame([{"OT": "III-1", "Categoría": "Desechos", "Cantidad": "1"}])

    excel_manager.send_data(
        filas, "Residuos", "Residuos", sp,
        headers_si_falta=_HEADERS_RESIDUOS_MIN,
    )

    wb2 = openpyxl.load_workbook(io.BytesIO(sp.uploaded))
    assert wb2.sheetnames.count("Residuos") == 1
    assert [c.value for c in wb2["Residuos"][1]] == _HEADERS_RESIDUOS_MIN
    assert [c.value for c in wb2["Residuos"][2]] == ["III-1", "Desechos", "1"]
    assert wb2["Residuos"].tables["Residuos"].ref == "A1:C2"
